import openpyxl as xl

wb = xl.load_workbook("transactions.xlsx")
sheet = wb['Sheet1']
cell = sheet['a1']
cell1 = sheet.cell(1,1)


for rows in range(2, sheet.max_row + 1):
    cell = sheet.cell(rows, 3)
    corrected_value = cell.value * 0.9
    corrected_value_cell = sheet.cell(rows, 4)
    corrected_value_cell.value = corrected_value
    fourth_cell_username = sheet.cell(1,4)
    fourth_cell_username.value = "Corrected Value"

    another_value = cell.value * 254
    fifth_cell_username = sheet.cell(1,5)
    fifth_cell_username.value = "Another Value"
    another_value_cell = sheet.cell(rows, 5)
    another_value_cell.value = another_value


wb.save('transactions2.xlsx')